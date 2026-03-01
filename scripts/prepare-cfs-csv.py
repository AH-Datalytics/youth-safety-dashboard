"""
Convert CFS Excel to per-sheet CSVs for fast TypeScript ETL.

The Calls for Service.xlsx file (577MB, 4.8M rows) is too large for
the Node.js xlsx library. This script converts each sheet to a CSV
that the TypeScript ETL can stream-parse quickly with papaparse.

Usage:
  python scripts/prepare-cfs-csv.py

Input:
  data/source/Calls for Service.xlsx

Output:
  data/generated/_cfs-sheets/Sheet1.csv, Sheet2.csv, ...
"""

import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    try:
        import pandas as pd
        USE_PANDAS = True
    except ImportError:
        print("Missing dependencies. Install with: pip install openpyxl (or pandas)")
        sys.exit(1)
    USE_PANDAS = False
else:
    USE_PANDAS = False

EXCEL_PATH = Path("data/source/Calls for Service.xlsx")
OUTPUT_DIR = Path("data/generated/_cfs-sheets")


def convert_with_pandas():
    import pandas as pd

    print(f"Reading {EXCEL_PATH} with pandas...")
    xls = pd.ExcelFile(EXCEL_PATH, engine="openpyxl")
    print(f"  Found {len(xls.sheet_names)} sheets: {', '.join(xls.sheet_names)}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    total_rows = 0
    for sheet_name in xls.sheet_names:
        print(f"  Converting sheet: {sheet_name}...")
        df = xls.parse(sheet_name)
        csv_path = OUTPUT_DIR / f"{sheet_name}.csv"
        df.to_csv(csv_path, index=False)
        total_rows += len(df)
        size_mb = os.path.getsize(csv_path) / (1024 * 1024)
        print(f"    {len(df):,} rows, {size_mb:.1f} MB")

    print(f"\nDone. {total_rows:,} total rows across {len(xls.sheet_names)} sheets")


def convert_with_openpyxl():
    import csv

    print(f"Reading {EXCEL_PATH} with openpyxl (read-only mode)...")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    print(f"  Found {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    total_rows = 0
    for sheet_name in wb.sheetnames:
        print(f"  Converting sheet: {sheet_name}...")
        ws = wb[sheet_name]
        csv_path = OUTPUT_DIR / f"{sheet_name}.csv"

        rows_written = 0
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
                rows_written += 1

        total_rows += rows_written
        size_mb = os.path.getsize(csv_path) / (1024 * 1024)
        print(f"    {rows_written:,} rows, {size_mb:.1f} MB")

    wb.close()
    print(f"\nDone. {total_rows:,} total rows across {len(wb.sheetnames)} sheets")


if __name__ == "__main__":
    if not EXCEL_PATH.exists():
        print(f"CFS Excel not found: {EXCEL_PATH}")
        print("Download source files first (scripts/download-sharepoint-files.ts)")
        sys.exit(0)  # Not an error — file might not exist in all environments

    try:
        import openpyxl
        convert_with_openpyxl()
    except ImportError:
        convert_with_pandas()
